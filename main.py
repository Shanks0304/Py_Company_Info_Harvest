import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from langchain.agents import initialize_agent, AgentType
from langchain_openai import ChatOpenAI
from langchain_core.output_parsers import JsonOutputParser
from config import OPENAI_API_KEY, INPUT_FILE, OUTPUT_FILE
from agent.tools import company_info_search
from agent.prompts import company_info_prompt, final_prompt
import re
import openpyxl

def read_company_names(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    company_names = [cell.value for cell in sheet['B'][1:] if cell.value]
    workbook.close()
    return company_names

def write_results_to_excel(results, file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Find the last column with data
    last_column = sheet.max_column
    
    # Write headers for new columns
    sheet.cell(row=1, column=last_column+1, value="Website")
    sheet.cell(row=1, column=last_column+2, value="Description of Products/Services")
    sheet.cell(row=1, column=last_column+3, value="Head office location (City, Country)")
    
    # Write results
    for result in results:
        company_name = result["company_name"]
        info = result["info"]
        
        # Find the row for this company
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == company_name:
                sheet.cell(row=row, column=3, value=info.get("Website", ""))
                sheet.cell(row=row, column=4, value=info.get("Description of Products/Services", ""))
                sheet.cell(row=row, column=5, value=info.get("Head office location (City, Country)", ""))
                break
    
    workbook.save(file_path)
    workbook.close()

def write_results(results, file_path):
    with open(file_path, 'w') as f:
        json.dump(results, f, indent=2)

def process_company(agent_executor, company_name):
    search_prompt = company_info_prompt.format(company_name=company_name)
    search_results = agent_executor.invoke(search_prompt)
    # Use the LLM directly to process the search results
    company_info_prompt_filled = final_prompt.format(
        company_name=company_name,
        search_results=search_results
    )
    company_info = agent_executor.invoke(company_info_prompt_filled)
    return company_info['output']

def json_parser(company_info):
    json_match = re.search(r'```json\s*(.*?)\s*```', company_info, re.DOTALL)
    if json_match:
        json_str = json_match.group(1)
        try:
            company_info_dict = json.loads(json_str)
            print(company_info_dict)
            return company_info_dict
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            print(f"Raw JSON string: {json_str}")
    else:
        print("No JSON found in the output")
        print(f"Raw output: {company_info['output']}")
    return None

def main():
    # Initialize the LLM
    llm = ChatOpenAI(temperature=0, model_name="gpt-4o", openai_api_key=OPENAI_API_KEY)

    # Initialize the agent
    agent = initialize_agent(
        [company_info_search],
        llm,
        agent=AgentType.ZERO_SHOT_REACT_DESCRIPTION,
        verbose=True
    )
    # Process companies
    company_names = read_company_names(INPUT_FILE)
    results = []

    # for company_name in company_names:
    #     try:
    #         company_info = process_company(agent, company_name)
    #         results.append({"company_name": company_name, "info": json_parser(company_info)})
            
    #         print(f"Processed: {company_name}")
    #     except Exception as e:
    #         print(f"Error processing {company_name}: {str(e)}")


    # Use ThreadPoolExecutor for concurrent processing
    with ThreadPoolExecutor(max_workers=3) as executor:
        future_to_company = {executor.submit(process_company, agent, company_name): company_name for company_name in company_names}
        
        for future in as_completed(future_to_company):
            company_name = future_to_company[future]
            try:
                company_info = future.result()
                results.append({"company_name": company_name, "info": json_parser(company_info)})
                print(f"Processed: {company_name}")
            except Exception as e:
                print(f"Error processing {company_name}: {str(e)}")

    write_results(results, OUTPUT_FILE)
    print(f"Processing complete. Results written to {OUTPUT_FILE}")
    write_results_to_excel(results, INPUT_FILE)

if __name__ == "__main__":
    main()